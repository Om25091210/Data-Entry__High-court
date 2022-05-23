import numpy as np
import pyperclip
import pandas as pd
from openpyxl import load_workbook
import pyrebase
from flask import Flask, render_template, request
import datetime
from firebase import Firebase
app = Flask(__name__)


firebaseConfig = {
  "apiKey": "AIzaSyBeG9CP59V5PJML0P2fzOxC4H_qD9ccoDw",
  "authDomain": "alertsystem-d55cb.firebaseapp.com",
  "databaseURL": "https://alertsystem-d55cb-default-rtdb.firebaseio.com",
  "projectId": "alertsystem-d55cb",
  "storageBucket": "alertsystem-d55cb.appspot.com",
  "messagingSenderId": "1059294088093",
  "appId": "1:1059294088093:web:ec6f920d6ceea00706fd74",
  "measurementId": "G-9JHDRQXHKJ",
  "serviceAccount": "service_files/service_account.json"
};
firebase = Firebase(firebaseConfig)
db = firebase.database()

config = {
  "apiKey": "AIzaSyBeG9CP59V5PJML0P2fzOxC4H_qD9ccoDw",
  "authDomain": "alertsystem-d55cb.firebaseapp.com",
  "databaseURL": "https://alertsystem-d55cb-default-rtdb.firebaseio.com",
  "storageBucket": "alertsystem-d55cb.appspot.com",
  "serviceAccount": "service_files/service_account.json"
}

firebase = pyrebase.initialize_app(config)

@app.route("/",methods = ['GET','POST'])
def index():

    #File must be uploaded as downloaded.xlsx

    storage = firebase.storage()
    storage.child("downloaded.xlsx").download("downloaded.xlsx")
    return render_template('index.html')

@app.route("/upload", methods=['POST'])
def upload():
    if request.method == 'POST':
        print("ok upload")
        storage = firebase.storage() #init firebase storage
        #upload back to cloud
        # as admin
        storage.child("downloaded.xlsx").put("downloaded.xlsx")
        
    return render_template('index.html');

@app.route("/download", methods=['POST'])
def download():
    if request.method == 'POST':
        print("ok download")
        storage = firebase.storage() #init firebase storage
        #upload back to cloud
        # as admin
        pyperclip.copy(storage.child("downloaded.xlsx").get_url(None))
       
        
    return render_template('index.html');

@app.route("/forward", methods=['POST'])
def move_forward():
    if request.method == 'POST':

        ps = request.form.get("ps")
        rmdate = request.form.get("rmdate")
        nd = request.form.get("nd")
        ct = request.form.get("ct")
        cn = request.form.get("cn")
        name = request.form.get("name")
        case_year = request.form.get("case_year")
        crime_no = request.form.get("crime_no")
        crime_year = request.form.get("crime_year")
        before = request.form.get("before")
        diary_date = request.form.get("diary_date")
        subject = request.form.get("subject")
            
        if(diary_date==""):
            diary_date="None"
        rm_Date_dot=datetime.datetime.strptime(rmdate, '%d/%m/%Y').strftime('%d.%m.%Y')
        before_dot=datetime.datetime.strptime(before, '%d/%m/%Y').strftime('%d.%m.%Y')
        fd_dot=datetime.datetime.strptime(rmdate, '%d/%m/%Y').strftime('%Y-%m-%d')
        fd=datetime.datetime.strptime(rmdate, '%d/%m/%Y').strftime('%Y.%m.%d').replace(".","")
        print(fd)
        key=fd+str(ps)+str(nd)+str(ct)+str(cn)+str(case_year)+str(crime_no)+str(crime_year);
        diction = {
                            'A':"",
                            'B':ps,
                            'C':nd,
                            'D':ct,
                            'E':cn,
                            'F':name,
                            'G':case_year,
                            'H':crime_no,
                            'I':crime_year,
                            'J':diary_date,
                            'K':rm_Date_dot,
                            'L':before_dot,
                            'date':fd_dot,
                            'pushkey':key,
                            'type':subject
                }
        db.child('data').child(key).set(diction)
        if(subject=="RM CALL"):
            
             alert = pd.read_excel('downloaded.xlsx',"RM CALL")
             sno=alert.shape[0]+2
             sno_insert=alert.shape[0]+1
             list_to_insert=[]
             wb = load_workbook("downloaded.xlsx")  # Work Book
             ws = wb["RM CALL"]
             for keys,values in diction.items():
                list_to_insert.append(values) #adding values from dictionary to a list
             list_to_insert[0]=str(sno_insert)
             for i in range(1,13):
                 ws.cell(row=sno, column=i).value = list_to_insert[i-1]
             wb.save('downloaded.xlsx')
        elif(subject=="RM RETURN"):

             alert = pd.read_excel('downloaded.xlsx',"RM RETURN")
             sno=alert.shape[0]+2
             sno_insert=alert.shape[0]+1
             list_to_insert=[]
             wb = load_workbook("downloaded.xlsx")  # Work Book
             ws = wb["RM RETURN"]
             for keys,values in diction.items():
                list_to_insert.append(values) #adding values from dictionary to a list
             list_to_insert[0]=str(sno_insert)
             for i in range(1,13):
                 ws.cell(row=sno, column=i).value = list_to_insert[i-1]
             wb.save('downloaded.xlsx')

        return render_template('index.html');

if __name__ == '__main__':
    app.run(debug=False,host='0.0.0.0')